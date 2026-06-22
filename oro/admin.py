from django.contrib import admin
from django.http import HttpResponse

# Register your models here.
from .models import Registration, teams
from import_export.admin import ExportMixin
from import_export import resources
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment



COMPETITION_COLORS = {
    'genius':  'FFD6E0',  # pink
    'robotex': 'D6EAF8',  # blue
    'ifest':   'D5F5E3',  # green
    'vex':     'FEF9E7',  # yellow
    'castic':  'F9EBEA',  # rose
    'jwp':     'E8DAEF',  # purple
    'ioai':    'FDEBD0',  # orange
    'isef':    'D6DBDF',  # greyvex
}


def export_colored_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    headers = ['Competition', 'Name', 'Email', 'Age', 'City', 'Club', 'ATASTian', 'Category', 'Project Title', 'Project Description', 'Team', 'Teammates']
    header_fill = PatternFill(fill_type='solid', fgColor='151515')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, reg in enumerate(queryset.order_by('competition'), 2):
        row_data = [
            reg.competition, reg.name, reg.mail, reg.age, reg.city,
            reg.club, 'Yes' if reg.atastian else 'No',
            reg.categorie, reg.project_title, reg.project_desc,
            'Yes' if reg.team else 'No', reg.teammates,
        ]
        color = COMPETITION_COLORS.get(reg.competition, 'FFFFFF')
        fill = PatternFill(fill_type='solid', fgColor=color)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="registrations.xlsx"'
    wb.save(response)
    return response

export_colored_excel.short_description = "Export selected as colored Excel"


class RegistrationResource(resources.ModelResource):
    class Meta:
        model = Registration

@admin.register(Registration)
class RegistrationAdmin(ExportMixin, admin.ModelAdmin):
    resource_classes = [RegistrationResource]
    actions = [export_colored_excel]
    list_display = ('name', 'mail', 'competition', 'categorie', 'club', 'city')
    list_filter = ('competition', 'categorie', 'club', 'atastian')
    search_fields = ('name', 'mail')
    ordering = ('competition',)


@admin.register(teams)
class TeamsAdmin(admin.ModelAdmin):
    list_display = ("team_name", "category", "competition", "score")
    list_filter = ("category", "competition")
    search_fields = ("team_name",)
    ordering = ("category", "-score", "team_name")
