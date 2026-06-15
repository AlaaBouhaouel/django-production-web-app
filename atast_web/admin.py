from django.contrib import admin
from django.http import HttpResponse
from django_summernote.admin import SummernoteModelAdmin
from import_export.admin import ExportMixin
from import_export import resources
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from .models import Registration, ifestImage, NewsArticle, Partners_Supporters, GalleryImage, Gifts, LatestNews


COMPETITION_COLORS = {
    'genius':  'FFD6E0',  # pink
    'robotex': 'D6EAF8',  # blue
    'ifest':   'D5F5E3',  # green
    'vex':     'FEF9E7',  # yellow
    'castic':  'F9EBEA',  # rose
    'jwp':     'E8DAEF',  # purple
    'ioai':    'FDEBD0',  # orange
    'isef':    'D6DBDF',  # grey
}


def export_colored_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    headers = ['Competition', 'Name', 'Email', 'Age', 'City', 'Club', 'ATASTian', 'Category', 'Project Title', 'Project Description', 'Team', 'Teammates']
    header_fill = PatternFill(fill_type='solid', fgColor='151515')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, reg in enumerate(queryset.order_by('competition'), 2):
        row_data = [
            reg.competition, reg.name, reg.mail, reg.age, reg.city,
            reg.club, 'Yes' if reg.atastian else 'No',
            reg.categorie, reg.project_title, reg.project_desc,
            'Yes' if reg.team else 'No', reg.teammates,
        ]
        color = COMPETITION_COLORS.get(reg.competition, 'FFFFFF')
        fill = PatternFill(fill_type='solid', fgColor=color)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="registrations.xlsx"'
    wb.save(response)
    return response

export_colored_excel.short_description = "Export selected as colored Excel"


@admin.register(ifestImage)
class ifestImageAdmin(admin.ModelAdmin):
    pass


class RegistrationResource(resources.ModelResource):
    class Meta:
        model = Registration

@admin.register(Registration)
class RegistrationAdmin(ExportMixin, admin.ModelAdmin):
    resource_classes = [RegistrationResource]
    actions = [export_colored_excel]
    list_display = ('name', 'mail', 'competition', 'categorie', 'club', 'city')
    list_filter = ('competition', 'categorie', 'club', 'atastian')
    search_fields = ('name', 'mail')
    ordering = ('competition',)


@admin.register(NewsArticle)
class NewsArticleAdmin(SummernoteModelAdmin):
    summernote_fields = ('body',)
    list_display = ('title', 'is_featured', 'published_date')
    list_filter = ('is_featured',)
    search_fields = ('title',)


@admin.register(Partners_Supporters)
class PartnersSupportersAdmin(admin.ModelAdmin):
    list_display = ('name', 'category', 'order')
    list_filter = ('category',)
    ordering = ('order',)


@admin.register(GalleryImage)
class GalleryImageAdmin(admin.ModelAdmin):
    list_display = ('__str__', 'category', 'uploaded_at')
    list_filter = ('category',)


@admin.register(Gifts)
class GiftsAdmin(admin.ModelAdmin):
    list_display = ('name', 'price', 'order')
    ordering = ('order',)


@admin.register(LatestNews)
class LatestNewsAdmin(admin.ModelAdmin):
    list_display = ('title', 'date')
    search_fields = ('title',)
    ordering = ('-date',)
